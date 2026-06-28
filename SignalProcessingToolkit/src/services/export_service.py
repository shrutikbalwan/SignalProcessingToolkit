from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

from src.models.fft_result import FFTResult
from src.models.project import Project
from src.models.signal import Signal
from src.plots.base import BasePlotWidget
from src.repositories.implementations.file_project_repository import FileProjectRepository
from src.repositories.implementations.file_signal_repository import FileSignalRepository


@dataclass
class ExportOptions:
    format: str = "csv"
    include_metadata: bool = True
    precision: int = 6
    dpi: int = 150


class ExportService:
    def __init__(self) -> None:
        self._signal_repo = FileSignalRepository()
        self._project_repo = FileProjectRepository()
        from src.plots.exporters import PlotExporter

        self._plot_exporter = PlotExporter()

    def export_signal(
        self, signal: Signal, path: Path, options: ExportOptions | None = None
    ) -> None:
        opts = options or ExportOptions()
        suffix = path.suffix.lower()
        if suffix == ".csv":
            self._signal_repo.export_csv(signal, path)
        elif suffix == ".txt":
            self._signal_repo.export_txt(signal, path)
        elif suffix == ".wav":
            self._signal_repo.export_wav(signal, path)
        elif suffix == ".npz":
            self._signal_repo.save(signal, path)
        elif suffix == ".xlsx":
            self._export_excel(signal, path, opts)
        else:
            raise ValueError(f"Unsupported format: {suffix}")

    def export_plot(self, plot_widget: BasePlotWidget, path: Path, dpi: int = 150) -> None:
        suffix = path.suffix.lower()
        if suffix == ".png":
            self._plot_exporter.export_png(plot_widget, path)
        elif suffix == ".svg":
            self._plot_exporter.export_svg(plot_widget, path)
        elif suffix == ".pdf":
            self._export_pdf(plot_widget, path, dpi)
        else:
            raise ValueError(f"Unsupported plot format: {suffix}")

    def export_report_pdf(
        self,
        signal: Signal | None,
        fft_result: FFTResult | None,
        plot_widgets: list[BasePlotWidget],
        path: Path,
        title: str = "Signal Processing Report",
    ) -> None:
        try:
            from fpdf import FPDF

            pdf = FPDF()
            pdf.add_page()
            pdf.set_font("Helvetica", size=16)
            pdf.cell(200, 10, text=title, new_x="LMARGIN", new_y="NEXT", align="C")
            pdf.ln(10)

            if signal is not None:
                pdf.set_font("Helvetica", size=12)
                pdf.cell(
                    200, 8, text=f"Signal: {signal.metadata.name}", new_x="LMARGIN", new_y="NEXT"
                )
                pdf.cell(
                    200, 8, text=f"Duration: {signal.duration:.4f}s", new_x="LMARGIN", new_y="NEXT"
                )
                pdf.cell(
                    200,
                    8,
                    text=f"Sampling Rate: {signal.sampling_rate:.0f} Hz",
                    new_x="LMARGIN",
                    new_y="NEXT",
                )
                pdf.cell(200, 8, text=f"RMS: {signal.rms:.4f}", new_x="LMARGIN", new_y="NEXT")
                pdf.ln(5)

            if fft_result is not None:
                pdf.set_font("Helvetica", size=12)
                pdf.cell(
                    200,
                    8,
                    text=f"Dominant Frequency: {fft_result.dominant_frequency:.2f} Hz",
                    new_x="LMARGIN",
                    new_y="NEXT",
                )
                pdf.cell(
                    200,
                    8,
                    text=f"Frequency Resolution: {fft_result.frequency_resolution:.2f} Hz",
                    new_x="LMARGIN",
                    new_y="NEXT",
                )
                pdf.ln(5)

            pdf.output(str(path))
        except ImportError:
            raise ImportError("fpdf2 is required for PDF report generation") from None

    def save_project(self, project: Project) -> None:
        self._project_repo.save(project)

    def load_project(self, path: Path) -> Project:
        return self._project_repo.load(path)

    def _export_excel(self, signal: Signal, path: Path, options: ExportOptions) -> None:
        try:
            from openpyxl import Workbook

            wb = Workbook()
            ws = wb.active
            ws.title = "Signal Data"
            ws.append(["time", "amplitude"])
            for t, y in zip(signal.time_vector, signal.time_data, strict=False):
                ws.append([round(t, options.precision), round(y, options.precision)])
            if options.include_metadata:
                ws2 = wb.create_sheet("Metadata")
                ws2.append(["Property", "Value"])
                ws2.append(["Name", signal.metadata.name])
                ws2.append(["Sampling Rate", signal.sampling_rate])
                ws2.append(["Frequency", signal.frequency])
                ws2.append(["Amplitude", signal.amplitude])
                ws2.append(["Duration", signal.duration])
            wb.save(str(path))
        except ImportError:
            raise ImportError("openpyxl is required for Excel export") from None

    def _export_pdf(self, plot_widget: BasePlotWidget, path: Path, dpi: int = 150) -> None:
        try:
            import matplotlib
            matplotlib.use("Agg")

            png_path = path.with_suffix(".png")
            self._plot_exporter.export_png(plot_widget, png_path)

            from fpdf import FPDF

            pdf = FPDF()
            pdf.add_page()
            pdf.image(str(png_path), x=10, y=10, w=180)
            pdf.output(str(path))
            png_path.unlink(missing_ok=True)
        except ImportError:
            raise ImportError("fpdf2 and matplotlib required for PDF plot export") from None
